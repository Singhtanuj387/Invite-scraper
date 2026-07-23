"""
WhatsApp Invite Scraper — Web Dashboard Backend
Flask application with SSE streaming for real-time monitoring.
"""

import os
import csv
import json
import time
import threading
from datetime import datetime
from queue import Queue

from flask import (
    Flask, render_template, request, jsonify,
    send_file, Response, stream_with_context
)

# Import the core scraper function
from scraper import check_whatsapp_invite

app = Flask(__name__)

# ---------------------------------------------------------------------------
# Directories
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

import uuid

# ---------------------------------------------------------------------------
# Shared state
# ---------------------------------------------------------------------------
jobs = {}
sse_subscribers = {}
sse_lock = threading.Lock()

def broadcast_event(job_id: str, event_type: str, data: dict):
    """Push an SSE event to every connected subscriber for a job."""
    payload = f"event: {event_type}\ndata: {json.dumps(data)}\n\n"
    dead = []
    with sse_lock:
        subs = sse_subscribers.get(job_id, [])
        for q in subs:
            try:
                q.put_nowait(payload)
            except Exception:
                dead.append(q)
        for q in dead:
            subs.remove(q)


# ---------------------------------------------------------------------------
# Background scraper
# ---------------------------------------------------------------------------
def run_scraper(job_id: str):
    """Run the scraper in a background thread, streaming results via SSE."""
    job = jobs.get(job_id)
    if not job:
        return

    fieldnames = [
        "Invite Link", "Group Name", "Status", "Members",
        "Country", "Admin", "Processing Status", "Timestamp of Scan", "Notes"
    ]

    # Read links from the uploaded CSV
    links = []
    try:
        with open(job["input_file"], mode="r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                if row:
                    link = row[0].strip()
                    if link.startswith("http"):
                        links.append(link)
    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)
        broadcast_event(job_id, "error", {"message": str(e)})
        return

    if not links:
        job["status"] = "error"
        job["error"] = "No valid links found in the uploaded CSV."
        broadcast_event(job_id, "error", {"message": job["error"]})
        return

    job["total"] = len(links)
    job["processed"] = 0
    job["success"] = 0
    job["failed"] = 0
    job["expired"] = 0
    job["results"] = []
    job["status"] = "running"
    job["started_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    job["finished_at"] = None
    job["error"] = None

    broadcast_event(job_id, "started", {
        "total": len(links),
        "started_at": job["started_at"],
    })

    try:
        with open(job["output_file"], mode="w", newline="", encoding="utf-8") as outfile:
            writer = csv.DictWriter(outfile, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
            writer.writeheader()

            for idx, link in enumerate(links, start=1):
                # Check if stop was requested
                if job["stop_event"].is_set():
                    job["status"] = "stopped"
                    broadcast_event(job_id, "stopped", {
                        "processed": job["processed"],
                        "total": job["total"],
                    })
                    return

                # Process the link
                result = check_whatsapp_invite(link)
                writer.writerow(result)
                outfile.flush()

                # Update counters
                job["processed"] = idx
                if result["Processing Status"] == "Success":
                    job["success"] += 1
                    if result["Status"] == "Expired":
                        job["expired"] += 1
                else:
                    job["failed"] += 1

                job["results"].append(result)
                if len(job["results"]) > 100:
                    job["results"].pop(0)

                # Broadcast the result
                broadcast_event(job_id, "result", {
                    "index": idx,
                    "total": len(links),
                    "row": result,
                    "counters": {
                        "processed": job["processed"],
                        "success": job["success"],
                        "failed": job["failed"],
                        "expired": job["expired"],
                        "total": job["total"],
                    }
                })

                # Rate-limit to be polite
                time.sleep(0.5)

    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)
        broadcast_event(job_id, "error", {"message": str(e)})
        return

    job["status"] = "done"
    job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    broadcast_event(job_id, "done", {
        "processed": job["processed"],
        "total": job["total"],
        "finished_at": job["finished_at"],
    })


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


def init_job():
    job_id = uuid.uuid4().hex
    jobs[job_id] = {
        "status": "idle",
        "total": 0,
        "processed": 0,
        "success": 0,
        "failed": 0,
        "expired": 0,
        "input_file": os.path.join(UPLOAD_DIR, f"input_{job_id}.csv"),
        "output_file": os.path.join(OUTPUT_DIR, f"output_{job_id}.csv"),
        "started_at": None,
        "finished_at": None,
        "error": None,
        "results": [],
        "stop_event": threading.Event()
    }
    sse_subscribers[job_id] = []
    return job_id, jobs[job_id]


@app.route("/api/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No file part in request"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    filename = file.filename.lower()
    job_id, job = init_job()
    save_path = job["input_file"]
    
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        import openpyxl
        from io import BytesIO
        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            ws = wb.active
            with open(save_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    if row and row[0]:
                        writer.writerow([str(row[0]).strip()])
        except Exception as e:
            return jsonify({"error": f"Error reading Excel file: {str(e)}"}), 400
    elif filename.endswith(".csv"):
        file.save(save_path)
    else:
        return jsonify({"error": "Only CSV and Excel files are accepted"}), 400

    # Count links
    count = 0
    with open(save_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if row and row[0].strip().startswith("http"):
                count += 1

    return jsonify({
        "success": True,
        "job_id": job_id,
        "filename": file.filename,
        "link_count": count,
    })


@app.route("/api/upload_text", methods=["POST"])
def upload_text():
    data = request.get_json(force=True) or {}
    text = data.get("text", "")
    if not text.strip():
        return jsonify({"error": "Empty text"}), 400
    
    links = []
    for line in text.split("\n"):
        line = line.strip()
        if line.startswith("http"):
            links.append(line)
            
    if not links:
        return jsonify({"error": "No valid http links found in text"}), 400
        
    job_id, job = init_job()
    save_path = job["input_file"]
    with open(save_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for link in links:
            writer.writerow([link])
            
    return jsonify({
        "success": True,
        "job_id": job_id,
        "filename": "Pasted Links",
        "link_count": len(links),
    })


@app.route("/api/start", methods=["POST"])
def start():
    job_id = request.args.get("job_id")
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Invalid job ID"}), 404
    if job["status"] == "running":
        return jsonify({"error": "Scraper is already running"}), 409

    job["stop_event"].clear()
    t = threading.Thread(target=run_scraper, args=(job_id,), daemon=True)
    t.start()
    return jsonify({"success": True, "message": "Scraper started"})


@app.route("/api/stop", methods=["POST"])
def stop():
    job_id = request.args.get("job_id")
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Invalid job ID"}), 404
    if job["status"] != "running":
        return jsonify({"error": "Scraper is not running"}), 409

    job["stop_event"].set()
    job["status"] = "stopping"
    return jsonify({"success": True, "message": "Stop signal sent"})


@app.route("/api/status", methods=["GET"])
def status():
    job_id = request.args.get("job_id")
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Invalid job ID"}), 404
    return jsonify({
        "status": job["status"],
        "total": job["total"],
        "processed": job["processed"],
        "success": job["success"],
        "failed": job["failed"],
        "expired": job["expired"],
        "started_at": job["started_at"],
        "finished_at": job["finished_at"],
        "error": job["error"],
    })


@app.route("/api/results", methods=["GET"])
def results():
    job_id = request.args.get("job_id")
    job = jobs.get(job_id)
    if not job: return jsonify([])
    return jsonify(job["results"])


@app.route("/api/stream")
def stream():
    job_id = request.args.get("job_id")
    q: Queue = Queue()
    with sse_lock:
        if job_id not in sse_subscribers:
            sse_subscribers[job_id] = []
        sse_subscribers[job_id].append(q)

    def event_stream():
        try:
            while True:
                payload = q.get()  # blocks until an event is available
                yield payload
        except GeneratorExit:
            pass
        finally:
            with sse_lock:
                if job_id in sse_subscribers and q in sse_subscribers[job_id]:
                    sse_subscribers[job_id].remove(q)

    return Response(
        stream_with_context(event_stream()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


def filter_row(row, filter_type):
    if not filter_type or filter_type == 'all':
        return True
    
    status = row.get("Status", "")
    proc_status = row.get("Processing Status", "")
    
    if filter_type == 'valid':
        return status == 'Active'
    elif filter_type == 'invalid':
        return status != 'Active'
    elif filter_type == 'expired':
        return status == 'Expired'
    elif filter_type == 'error':
        return proc_status == 'Failed'
    return True


@app.route("/api/download", methods=["GET"])
def download():
    job_id = request.args.get("job_id")
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Invalid job ID"}), 404

    filter_type = request.args.get("filter", "all")
    is_batch = request.args.get("batch", "false").lower() == "true"
    
    output_path = job["output_file"]
    if not output_path or not os.path.exists(output_path):
        return jsonify({"error": "No output file available yet"}), 404

    temp_dir = os.path.join(OUTPUT_DIR, f"temp_{job_id}_{int(time.time())}")
    os.makedirs(temp_dir, exist_ok=True)
    import zipfile
    
    if is_batch:
        zip_path = os.path.join(OUTPUT_DIR, f"scraper_output_{filter_type}_csv.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            with open(output_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames
                
                batch_num = 1
                row_count = 0
                current_temp = os.path.join(temp_dir, f"batch_{batch_num}.csv")
                out_f = open(current_temp, "w", encoding="utf-8", newline="")
                writer = csv.DictWriter(out_f, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
                writer.writeheader()
                
                for row in reader:
                    if filter_row(row, filter_type):
                        writer.writerow(row)
                        row_count += 1
                        if row_count >= 1000:
                            out_f.close()
                            zf.write(current_temp, f"scraper_output_{filter_type}_batch_{batch_num}.csv")
                            batch_num += 1
                            row_count = 0
                            current_temp = os.path.join(temp_dir, f"batch_{batch_num}.csv")
                            out_f = open(current_temp, "w", encoding="utf-8", newline="")
                            writer = csv.DictWriter(out_f, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
                            writer.writeheader()
                
                out_f.close()
                if row_count > 0:
                    zf.write(current_temp, f"scraper_output_{filter_type}_batch_{batch_num}.csv")
                    
        return send_file(zip_path, as_attachment=True, download_name=f"scraper_output_{filter_type}_batches.zip")
        
    else:
        if filter_type == "all":
            return send_file(output_path, mimetype="text/csv", as_attachment=True, download_name="scraper_output.csv")
            
        temp_csv = os.path.join(temp_dir, "filtered.csv")
        with open(output_path, "r", encoding="utf-8") as f, open(temp_csv, "w", encoding="utf-8", newline="") as out_f:
            reader = csv.DictReader(f)
            writer = csv.DictWriter(out_f, fieldnames=reader.fieldnames, quoting=csv.QUOTE_ALL)
            writer.writeheader()
            for row in reader:
                if filter_row(row, filter_type):
                    writer.writerow(row)
        return send_file(temp_csv, mimetype="text/csv", as_attachment=True, download_name=f"scraper_output_{filter_type}.csv")


@app.route("/api/download/excel", methods=["GET"])
def download_excel():
    job_id = request.args.get("job_id")
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Invalid job ID"}), 404

    filter_type = request.args.get("filter", "all")
    is_batch = request.args.get("batch", "false").lower() == "true"
    
    output_path = job["output_file"]
    if not output_path or not os.path.exists(output_path):
        return jsonify({"error": "No output file available yet"}), 404

    import openpyxl
    import zipfile
    temp_dir = os.path.join(OUTPUT_DIR, f"temp_{job_id}_{int(time.time())}")
    os.makedirs(temp_dir, exist_ok=True)
    
    if is_batch:
        zip_path = os.path.join(OUTPUT_DIR, f"scraper_output_{filter_type}_excel.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            with open(output_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames
                
                batch_num = 1
                row_count = 0
                
                wb = openpyxl.Workbook(write_only=True)
                ws = wb.create_sheet()
                ws.append(fieldnames)
                
                for row in reader:
                    if filter_row(row, filter_type):
                        ws.append(list(row.values()))
                        row_count += 1
                        if row_count >= 1000:
                            temp_excel = os.path.join(temp_dir, f"batch_{batch_num}.xlsx")
                            wb.save(temp_excel)
                            zf.write(temp_excel, f"scraper_output_{filter_type}_batch_{batch_num}.xlsx")
                            batch_num += 1
                            row_count = 0
                            wb = openpyxl.Workbook(write_only=True)
                            ws = wb.create_sheet()
                            ws.append(fieldnames)
                
                if row_count > 0:
                    temp_excel = os.path.join(temp_dir, f"batch_{batch_num}.xlsx")
                    wb.save(temp_excel)
                    zf.write(temp_excel, f"scraper_output_{filter_type}_batch_{batch_num}.xlsx")
                    
        return send_file(zip_path, as_attachment=True, download_name=f"scraper_output_{filter_type}_batches.zip")
        
    else:
        temp_excel = os.path.join(temp_dir, "filtered.xlsx")
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        
        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            ws.append(reader.fieldnames)
            for row in reader:
                if filter_row(row, filter_type):
                    ws.append(list(row.values()))
                    
        wb.save(temp_excel)
        return send_file(
            temp_excel,
            as_attachment=True,
            download_name=f"scraper_output_{filter_type}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True, threaded=True)
